
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

places = [
    # [Name, Area, Address, Lat, Lng]
    ["Gài Yâang Rá Bìiap (grilled chicken and som tum)", "Khon Kaen", "391/5 Thepharak Alley, Tambon Nai Mueang, Mueang Khon Kaen District, Khon Kaen 40000", "16.441827", "102.830819"],
    ["แก่น I KAEN", "Khon Kaen", "140 ซ. อดุลยาราม 3, Nai Mueang, Mueang Khon Kaen District, Khon Kaen 40000", "16.463885", "102.825736"],
    ["Baan Heng", "Khon Kaen", "", "16.434136", "102.835820"],
    ["Ton Tann Market", "Khon Kaen", "", "16.418626", "102.817944"],
    ["Prasit Pochana", "Khon Kaen", "4 Na Mueang Rd, Nai Mueang, Mueang Khon Kaen District, Khon Kaen 40000", "16.434253", "102.833379"],
    ["Aim Oat", "Khon Kaen", "", "16.435502", "102.836260"],
    ["ประตูมอดินแดง (Khon Kaen University Gate)", "Khon Kaen", "123 Mittraphap Rd, Nai Mueang, Mueang Khon Kaen 40001", "16.476060", "102.832527"],
    ["GaGa Udon House", "Bangkok", "No. 6/5, H Building, 1st & 2nd Floor, Phrom Si 2 Alley, Khlong Tan Nuea, Watthana, Bangkok 10110", "13.736441", "100.574091"],
    ["CHAGÔ ชาโก้ - EmQuartier", "Bangkok", "Sukhumvit Rd, Khlong Tan Nuea, Watthana, Bangkok 10110", "13.730498", "100.569468"],
    ["Franca | Modern Italian and Fine Steaks", "Bangkok", "Sukhumvit Rd, Khlong Tan Nuea, Watthana, Bangkok 10110", "13.730498", "100.569468"],
    ["Sourcream Factory @ PARC Bangna", "Bangkok", "2/3 Moo 14, Bangna-Trad Rd, Bang Kaeo, Bang Phli District, Samut Prakan 10540", "13.671000", "100.730000"],
    ["Royal Osha", "Bangkok", "Sukhumvit Rd, Khlong Tan Nuea, Watthana, Bangkok", "13.745000", "100.560000"],
    ["Owl cafe Bangkok", "Bangkok", "Soi Ari, Phahon Yothin Rd, Sam Sen Nai, Phaya Thai, Bangkok", "13.772000", "100.545000"],
    ["The Cafe Mookata Ari", "Bangkok", "Phahon Yothin Rd, Sam Sen Nai, Phaya Thai, Bangkok", "13.772000", "100.545000"],
    ["EmQuartier", "Bangkok", "693 Sukhumvit Rd, Khlong Tan Nuea, Watthana, Bangkok 10110", "13.730498", "100.569468"],
    ["ทิพย์วารี (convenience store)", "Bangkok", "Sukhumvit area, Bangkok", "13.730000", "100.570000"],
    ["Smith & Co. (Chidlom)", "Bangkok", "Chidlom, Pathum Wan, Bangkok", "13.745000", "100.545000"],
    ["Butterbear Cafe Siam Paragon", "Bangkok", "991 Rama I Rd, Pathum Wan, Bangkok 10330", "13.745800", "100.534800"],
    ["Butterbear Cafe EmSphere", "Bangkok", "Sukhumvit Rd, Khlong Tan, Khlong Toei, Bangkok 10110", "13.720000", "100.575000"],
    ["Paragon Department Store", "Bangkok", "991 Rama I Rd, Pathum Wan, Bangkok 10330", "13.745800", "100.534800"],
    ["Toscana Valley", "Khao Yai", "Pong Ta Long, Pak Chong District, Nakhon Ratchasima 30130", "14.517816", "101.507436"],
    ["Huen Muan Jai", "Chiang Mai", "24 Ratchaphuek Alley, Chang Phueak, Mueang Chiang Mai District, Chiang Mai 50300", "18.799861", "98.975599"],
    ["Ngon Restaurant", "Vietnam", "160 Pasteur, Bến Nghé, Quận 1, Ho Chi Minh City, Vietnam", "10.777481", "106.699388"],
    ["kokoa - ice cream & coffee artisans (SALA)", "Koh Samui", "", "9.537903", "100.069523"],
    ["Koh Samui pin", "Koh Samui", "9.5328889, 100.0689444", "9.532889", "100.068944"],
    ["The Tent", "Koh Samui", "", "9.537254", "100.070196"],
    ["The Jungle Club", "Koh Samui", "", "9.499578", "100.047052"],
    ["Sabienglae, Bang Rak Beach", "Koh Samui", "", "9.558256", "100.055883"],
    ["SALA Samui Chaweng Beach", "Koh Samui", "", "9.537782", "100.069772"],
    ["Chin Bo Dang - Central World", "Bangkok", "999/9 Rama I Rd, Pathum Wan, Bangkok 10330", "13.746198", "100.539324"],
    ["Chin Bo Dang", "Bangkok", "693 Sukhumvit Rd, Khlong Tan Nuea, Watthana, Bangkok 10110", "13.730498", "100.569468"],
    ["SUSHI SEKI (Siam Paragon)", "Bangkok", "991 Rama I Rd, Pathum Wan, Bangkok 10330", "13.745800", "100.534800"],
    ["KAEW BOUTIQUE", "Bangkok", "Bangkok", "13.745000", "100.535000"],
    ["Pan Pan Italian Restaurant", "Pattaya", "313 Muang Pattaya, Bang Lamung District, Chon Buri 20150", "12.905064", "100.868955"],
    ["บังดีนไก่หาดใหญ่", "Bangkok", "Saen Suk Alley, Khlong Tan, Khlong Toei, Bangkok 10110", "13.717730", "100.572849"],
    ["chago", "Bangkok", "Siam area, Bangkok", "13.744559", "100.533533"],
    ["Ternajachob cafe", "Bangkok", "6 Chaloem Phrakiat Ratchakan Thi 9 Rd, Prawet, Bangkok 10250", "13.708106", "100.692569"],
    ["TumLubThai Khanom Krok", "Bangkok", "Soi Wanit 2, Talat Noi, Samphanthawong, Bangkok 10100", "13.732052", "100.513641"],
    ["Eak-O-Cha stewed beef", "Bangkok", "3 Soi Lat Phrao 71, Khlong Chaokhun Sing, Wang Thonglang, Bangkok 10310", "13.794076", "100.608867"],
    ["Ton Yen Ta Four", "Bangkok", "89 3 Soi Lat Phrao 71, Saphan Song, Wang Thonglang, Bangkok 10310", "13.793747", "100.608791"],
    ["WANWAN coffee roasters", "Bangkok", "Nak Niwat Rd, Lat Phrao, Bangkok 10230", "13.805569", "100.607499"],
    ["ร้านก๋วยเตี๋ยวเนื้อวัวนายเที๊ยบ สาขานาคนิวาส 10", "Bangkok", "159 Soi Lat Phrao 71, Lat Phrao, Bangkok 10230", "13.803895", "100.606375"],
    ["Kanom-Jaoka X ZEPHYR", "Bangkok", "28 Soi Mu Ban Suan Non Ville, Bang Rak Noi, Mueang Nonthaburi 11000", "13.879828", "100.458708"],
    ["CHANN | Bangkok-Noi", "Bangkok", "30/1 Soi Somdej Pra Pinklaow 1, Arun Amarin, Bangkok Noi, Bangkok 10700", "13.761466", "100.486838"],
    ["Shibainu Story (Farm & Cafe)", "Bangkok", "188/2 Soi Chulalongkorn 5, Wang Mai, Pathum Wan, Bangkok 10330", "13.741847", "100.524093"],
    ["Shell Asoke V-Power Station", "Bangkok", "Asok Montri Rd, Khlong Toei Nuea, Watthana, Bangkok 10110", "13.743134", "100.562412"],
    ["Such A Small World", "Bangkok", "951 ถ. เจริญกรุง, Talat Noi, Samphanthawong, Bangkok 10100", "13.732060", "100.514999"],
    ["Prannok", "Bangkok", "Siriraj, Bangkok Noi, Bangkok 10700", "13.755811", "100.486813"],
    ["Café Amazon (Wang Lang Soi 1)", "Bangkok", "339 Arun Amarin Rd, Siriraj, Bangkok Noi, Bangkok 10700", "13.756240", "100.486280"],
    ["Kanom Krok Wang Lang (Thai Coconut Pancake)", "Bangkok", "Wang Lang Market, Pranok Rd, Siriraj, Bangkok Noi, Bangkok 10700", "13.756094", "100.485806"],
    ["KAO PIAK SEN BKK (Gaysorn Amarin)", "Bangkok", "G Floor, Gaysorn Amarin, Ratchadamri Rd, Pathum Wan, Bangkok", "13.745000", "100.540000"],
    ["Somtam Nua", "Bangkok", "392/14 Siam Square Soi 5, Pathum Wan, Bangkok 10330", "13.745562", "100.532697"],
    ["Yua Cafe & Dining", "Bangkok", "Bangkok", "13.745000", "100.535000"],
    ["Kolun.h", "Bangkok", "Bangkok", "13.745000", "100.565000"],
    ["Yoru Omakase", "Bangkok", "Bangkok", "13.745000", "100.565000"],
    ["Sushi Ichizu 鮨いちづ", "Bangkok", "Bangkok", "13.745000", "100.565000"],
    ["Shiro-i Omakase", "Bangkok", "Bangkok", "13.745000", "100.565000"],
    ["Davin Cafe", "Bangkok", "Bangkok", "13.745000", "100.535000"],
    ["Ekamai Mookata", "Bangkok", "Ekamai, Watthana, Bangkok", "13.718000", "100.585000"],
    ["Thea Cafe", "Bangkok", "Bangkok", "13.745000", "100.565000"],
    ["เกศเตี๋ยว ก๋วยเตี๋ยวเรือ (Kate Teaw Boat Noodles Siam Square Soi 3)", "Bangkok", "Siam Square Soi 3, Pathum Wan, Bangkok", "13.744951", "100.532884"],
    ["Jae Wan", "Bangkok", "Pathum Wan, Bangkok", "13.739649", "100.522226"],
    ["Jurassic World: The Experience Bangkok", "Bangkok", "Bangkok", "13.704360", "100.502191"],
    ["ข้ามันบ้านนอก by บ้านนอกคอกนาเขาใหญ่", "Bangkok", "Bangkok (Lat Krabang area)", "13.772791", "100.661015"],
    ["Sam Lor", "Bangkok", "1076 ถ. เจริญกรุง, Bang Rak, Bangkok 10500", "13.730250", "100.515556"],
    ["漢王廟 (Han Wang Miao)", "Bangkok", "1192 Soi Chareonkrung 22, Talat Noi, Samphanthawong, Bangkok 10100", "13.732880", "100.512180"],
    ["Galaxy Enterprise Gaming Thailand", "Bangkok", "4th floor Zone A, Center 444, Wang Mai, Pathum Wan, Bangkok 10330", "13.744714", "100.529636"],
    ["王朗市場 (Wang Lang Market)", "Bangkok", "13 Soi Wang Lang, Siriraj, Bangkok Noi, Bangkok 10700", "13.756087", "100.486287"],
    ["迪奧咖啡 曼谷 (Dior Cafe Bangkok)", "Bangkok", "1029 Phloen Chit Rd, Lumphini, Pathum Wan, Bangkok 10330", "13.744155", "100.545199"],
    ["IMPACT Arena", "Bangkok", "Popular Rd, Ban Mai, Pak Kret District, Nonthaburi 11120", "13.912606", "100.547753"],
    ["Tue Kha Tang Stewed Pork Leg Jelly", "Bangkok", "Talat Noi, Samphanthawong, Bangkok", "13.735293", "100.512145"],
    ["Hong Sieng Kong", "Bangkok", "734-736 Soi Wanit 2, Talat Noi, Samphanthawong, Bangkok 10100", "13.734795", "100.511591"],
    ["32Bar X", "Bangkok", "294 Soi Wanit 2, Talat Noi, Samphanthawong, Bangkok 10100", "13.733919", "100.512114"],
    ["Jay Hieng Fishball Noodles", "Bangkok", "855 Soi Wanit 2, Talat Noi, Samphanthawong, Bangkok 10100", "13.733477", "100.513496"],
    ["Dear December Cafe", "Bangkok", "7 Soi Srinagarindra 58, Nong Bon, Prawet, Bangkok 10250", "13.675657", "100.644664"],
    ["Maleenont Tower (CH3)", "Bangkok", "3199 Rama IV Rd, Khlong Tan, Khlong Toei, Bangkok 10110", "13.717049", "100.572504"],
    ["Kwong Kee Roast - 鄺記", "Bangkok", "Charoen Krung Rd, Bang Rak, Bangkok", "13.732000", "100.515000"],
]

# Area color mapping
AREA_COLORS = {
    "Bangkok": "FFF2CC",       # light yellow
    "Khon Kaen": "DAE8FC",    # light blue
    "Koh Samui": "D5E8D4",    # light green
    "Chiang Mai": "E1D5E7",   # light purple
    "Khao Yai": "FFE6CC",     # light orange
    "Pattaya": "FFD966",      # amber
    "Vietnam": "F8CECC",      # light red
}

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Locations"

# Header
headers = ["#", "Name", "Area", "Address", "Latitude", "Longitude", "Google Maps"]
header_fill = PatternFill("solid", start_color="2F4F4F")
header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border

ws.row_dimensions[1].height = 30

# Data rows
seen = set()
row_num = 2
written = 0
for place in places:
    name = place[0]
    if name in seen:
        continue
    seen.add(name)

    area, address, lat, lng = place[1], place[2], place[3], place[4]
    maps_url = f"https://www.google.com/maps/search/{name.replace(' ', '+')}/@{lat},{lng},17z" if lat else ""

    row_data = [written + 1, name, area, address, float(lat) if lat else "", float(lng) if lng else "", maps_url]
    fill_color = AREA_COLORS.get(area, "FFFFFF")
    fill = PatternFill("solid", start_color=fill_color)

    for col, val in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.font = Font(name="Arial", size=10)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=(col == 2 or col == 4))
        if col in (5, 6):  # lat/lng
            cell.number_format = "0.000000"
        if col == 7 and val:  # URL
            cell.hyperlink = val
            cell.value = "📍 Map"
            cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")

    ws.row_dimensions[row_num].height = 20
    row_num += 1
    written += 1

# Column widths
ws.column_dimensions["A"].width = 5   # #
ws.column_dimensions["B"].width = 42  # Name
ws.column_dimensions["C"].width = 14  # Area
ws.column_dimensions["D"].width = 48  # Address
ws.column_dimensions["E"].width = 12  # Lat
ws.column_dimensions["F"].width = 12  # Lng
ws.column_dimensions["G"].width = 10  # Maps link

# Freeze header
ws.freeze_panes = "A2"

# Add a legend sheet
ls = wb.create_sheet("Legend")
ls["A1"] = "Area Color Legend"
ls["A1"].font = Font(bold=True, name="Arial", size=12)
ls.column_dimensions["A"].width = 20
ls.column_dimensions["B"].width = 20
for i, (area, color) in enumerate(AREA_COLORS.items(), 2):
    ls.cell(row=i, column=1, value=area).font = Font(name="Arial", size=10)
    ls.cell(row=i, column=1).fill = PatternFill("solid", start_color=color)
    ls.cell(row=i, column=2, value=f"{written} total (see Locations sheet)").font = Font(name="Arial", size=10, color="888888")

out = "/sessions/elegant-happy-lamport/mnt/lingorm_bangkok_map/Lingorm_Bangkok_Locations.xlsx"
wb.save(out)
print(f"Saved {written} locations to {out}")
